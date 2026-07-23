# Django Imports
from django.conf import settings
from django.contrib import messages
from django.http import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404

# Python Imports
import os
from collections import defaultdict
from datetime import datetime, timedelta

# Excel
from openpyxl import Workbook, load_workbook

# Forms
from .forms import JobApplicationForm

# Models
from .models import (
    Career,
    Certificate,
    ClientProject,
    Contact,
    Course,
    CourseBooking,
    DemoCategory,
    DemoRequest,
    Internship,
    Placement,
    ProcessStep,
    Service,
    ServiceBooking,
    StudentReview,
    UpcomingWorkshop,
    WorkshopPhoto,
    WorkshopRegistration,
)

# Brevo Email Helper
from .brevo_mail import send_brevo_email


def home(request):
    if request.method == "POST":
        Contact.objects.create(
            name=request.POST.get("name"),
            email=request.POST.get("email"),
            message=request.POST.get("message"),
        )
        return redirect("home")

    context = {
        "courses": Course.objects.all(),
        "services": Service.objects.all(),
        "projects": ClientProject.objects.all(),
        "reviews": StudentReview.objects.all(),
        "total_bookings": CourseBooking.objects.count(),
        "workshops": WorkshopPhoto.objects.all(),
        "certificates": Certificate.objects.all(),
        "internships": Internship.objects.all(),
        "upcoming_workshop": UpcomingWorkshop.objects.first(),
    }

    return render(request, "index.html", context)


def review_page(request):
    reviews = StudentReview.objects.all().order_by("-created_at")
    return render(request, "reviews.html", {"reviews": reviews})


def course_detail(request, id):
    course = get_object_or_404(Course, id=id)

    context = {
        "course": course,
        "description_lines": course.description.splitlines(),
    }

    return render(request, "courses.html", context)


def allcourse(request):
    context = {
        "allcourse": Course.objects.all(),
    }
    return render(request, "allcourse.html", context)


def about(request):
    return render(request, "about.html")


def courses(request):
    context = {
        "courses": Course.objects.all(),
    }
    return render(request, "courses.html", context)


def services(request):
    context = {
        "services": Service.objects.all(),
    }
    return render(request, "services.html", context)


def service_detail(request, pk):
    service = get_object_or_404(Service, pk=pk)

    context = {
        "service": service,
        "featured_demos": service.demo_links.filter(is_featured=True),
        "features": service.features.all(),
        "faqs": service.faqs.all(),
        "process_steps": ProcessStep.objects.all(),
        "demo_categories": DemoCategory.objects.filter(is_active=True),
    }

    return render(request, "service_detail.html", context)


def contact(request):
    if request.method == "POST":
        Contact.objects.create(
            name=request.POST.get("name"),
            email=request.POST.get("email"),
            message=request.POST.get("message"),
        )
        return redirect("contact")

    return render(request, "contact.html")


def placement(request):
    context = {
        "placements": Placement.objects.all(),
    }

    return render(request, "placement.html", context)


def career_list(request):
    careers = Career.objects.all().order_by("-posted_on")
    return render(request, "career.html", {"careers": careers})


def career_detail(request, pk):
    career = get_object_or_404(Career, pk=pk)
    return render(request, "career_detail.html", {"career": career})


def apply_job(request, pk):
    career = get_object_or_404(Career, pk=pk)

    if request.method == "POST":
        form = JobApplicationForm(request.POST, request.FILES)

        if form.is_valid():
            application = form.save(commit=False)
            application.career = career
            application.save()

            # ================= SAVE TO EXCEL =================

            job_folder = os.path.join(settings.MEDIA_ROOT, "job_applications")
            os.makedirs(job_folder, exist_ok=True)

            file_path = os.path.join(job_folder, "job_applications.xlsx")

            if os.path.exists(file_path):
                workbook = load_workbook(file_path)
                sheet = workbook.active
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.append([
                    "Full Name",
                    "Email",
                    "Phone",
                    "Role Applied",
                    "Cover Letter",
                    "Resume File Name",
                    "Applied Date",
                ])

            sheet.append([
                application.full_name,
                application.email,
                application.mobile,
                career.role,
                application.cover_letter,
                application.resume.name,
                datetime.now().strftime("%d-%m-%Y %H:%M"),
            ])

            workbook.save(file_path)

            # ================= EMAIL TO CANDIDATE =================

            candidate_html = f"""
            <p>Hi {application.full_name},</p>

            <p>Greetings from Errors2Experts!</p>

            <p>
            We have received your application for the role of
            <b>{career.role}</b>.
            </p>

            <p>
            Our HR team will review your profile and contact you soon.
            </p>

            <p>Thank you for your interest.</p>

            <br>

            <p>
            Regards,<br>
            Errors2Experts Team
            </p>
            """

            send_brevo_email(
                application.email,
                "Application Received - Errors2Experts",
                candidate_html,
            )

            # ================= EMAIL TO ADMIN =================

            admin_subject = f"🚀 New Job Application - {career.role}"

            admin_html = f"""
            <h2>New Job Application Received!</h2>

            <p><b>Candidate Details</b></p>

            <p><b>Name:</b> {application.full_name}</p>

            <p><b>Email:</b> {application.email}</p>

            <p><b>Phone:</b> {application.mobile}</p>

            <p><b>Role Applied:</b> {career.role}</p>

            <p><b>Cover Letter:</b></p>

            <p>{application.cover_letter}</p>

            <p>
            <b>Applied On:</b>
            {datetime.now().strftime("%d-%m-%Y %H:%M")}
            </p>

            <p>Please review the attached resume.</p>

            <br>

            <b>Errors2Experts System</b>
            """

            attachments = []

            if application.resume:
                attachments.append(application.resume.path)

            send_brevo_email(
                settings.ADMIN_NOTIFICATION_EMAIL,
                admin_subject,
                admin_html,
                attachments=attachments,
            )

            return render(request, "application_success.html")

    else:
        form = JobApplicationForm()

    return render(
        request,
        "apply_job.html",
        {
            "form": form,
            "career": career,
        },
    )


def demo_booking(request):
    if request.method == "POST":
        name = request.POST.get("name")
        email = request.POST.get("email")
        mobile = request.POST.get("mobile")
        education = request.POST.get("education")
        source = request.POST.get("source")

        # ===============================
        # SAVE TO EXCEL
        # ===============================

        file_path = os.path.join(settings.BASE_DIR, "demo_booking1.xlsx")

        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Demo Bookings"
            ws.append([
                "Name",
                "Email",
                "Mobile",
                "Education",
                "Source"
            ])

        ws.append([
            name,
            email,
            mobile,
            education,
            source
        ])

        wb.save(file_path)

        # ===============================
        # EMAIL TO ADMIN
        # ===============================

        admin_subject = "New Demo Booking - Errors2Experts"

        admin_html = f"""
        <h2>📩 New Demo Booking Received</h2>

        <p><strong>Name:</strong> {name}</p>
        <p><strong>Email:</strong> {email}</p>
        <p><strong>Mobile:</strong> {mobile}</p>
        <p><strong>Education:</strong> {education}</p>
        <p><strong>Source:</strong> {source}</p>
        """

        send_brevo_email(
            settings.ADMIN_NOTIFICATION_EMAIL,
            admin_subject,
            admin_html,
        )

        # ===============================
        # EMAIL TO USER
        # ===============================

        user_subject = "Welcome to Errors2Experts 🚀"

        user_html = f"""
        <h2>Hello {name},</h2>

        <p>Greetings from <strong>Errors2Experts!</strong></p>

        <p>
        Thank you for showing interest in learning with us.
        We have received your demo booking successfully.
        </p>

        <p>
        Our team will contact you soon and guide you through the next steps.
        </p>

        <p>
        At Errors2Experts, we believe:<br>
        <i>"Learn from errors, grow with knowledge, and become an expert."</i>
        </p>

        <h3>Stay Connected</h3>

        <p>
        🌐 Website: https://yourwebsite.com<br>
        📸 Instagram:
        https://www.instagram.com/errors2experts_2026/
        </p>

        <p>
        Let's learn, grow, and build your future together!
        </p>

        <br>

        <p>
        Best Regards,<br>
        <strong>Team Errors2Experts</strong>
        </p>
        """

        send_brevo_email(
            email,
            user_subject,
            user_html,
        )

        return JsonResponse({"status": "success"})

    return JsonResponse({"status": "failed"})


def book_course(request, id):
    course = get_object_or_404(Course, id=id)

    if request.method == "POST":

        name = request.POST.get("name")
        email = request.POST.get("email")
        mobile = request.POST.get("mobile")
        education = request.POST.get("education")
        year_passed = request.POST.get("year_passed")
        payment_type = "full"

        price = int(course.price)

        # ---------------- PAYMENT CALCULATION ----------------

        if payment_type == "emi":
            amount = price // 2
            balance = price - amount
            due_date = datetime.now() + timedelta(days=15)
        else:
            amount = price
            balance = 0
            due_date = None

        # ---------------- SAVE TO DATABASE ----------------

        CourseBooking.objects.create(
            name=name,
            email=email,
            mobile=mobile,
            education=education,
            year_passed=year_passed,
            course=course.title,
            amount=amount,
        )

        # ---------------- SAVE TO EXCEL ----------------

        booking_folder = os.path.join(settings.BASE_DIR, "media", "bookings")

        if not os.path.exists(booking_folder):
            os.makedirs(booking_folder)

        file_path = os.path.join(booking_folder, "course_bookings.xlsx")

        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append([
                "Name",
                "Email",
                "Mobile",
                "Education",
                "Year Passed",
                "Course",
                "Payment Type",
                "First Payment",
                "Balance",
                "Due Date",
                "Booking Date"
            ])

        sheet.append([
            name,
            email,
            mobile,
            education,
            year_passed,
            course.title,
            payment_type,
            amount,
            balance,
            due_date.strftime("%d-%m-%Y") if due_date else "N/A",
            datetime.now().strftime("%d-%m-%Y %H:%M")
        ])

        workbook.save(file_path)

        # ---------------- STUDENT EMAIL CONTENT ----------------

        if payment_type == "emi":
            payment_message = f"""
            <p><strong>Your First Payment:</strong> ₹{amount}</p>
            <p><strong>Remaining Balance:</strong> ₹{balance}</p>
            <p><strong>Next Payment Due Date:</strong> {due_date.strftime('%d-%m-%Y')}</p>

            <p>
            Kindly complete your first payment and send the screenshot via:
            </p>

            <p>
            WhatsApp: +91 9363342646<br>
            Email: errors2experts.official@gmail.com
            </p>

            <p><strong>Invoice is mandatory.</strong></p>
            """
        else:
            payment_message = f"""
            <p><strong>Total Amount:</strong> ₹{amount}</p>

            <p>
            Kindly complete your payment and send the screenshot via:
            </p>

            <p>
            WhatsApp: +91 9363342646<br>
            Email: errors2experts.official@gmail.com
            </p>

            <p><strong>Invoice is mandatory.</strong></p>
            """

        student_html = f"""
        <h2>Welcome to Errors2Experts!</h2>

        <p>Hi <strong>{name}</strong>,</p>

        <p>Thank you for booking your course with us.</p>

        <p><strong>Course:</strong> {course.title}</p>

        {payment_message}

        <hr>

        <h3>Terms & Conditions</h3>

        <ol>
            <li>Errors2Experts is a MSME registered training institute.</li>
            <li>Amount once paid is non-refundable.</li>
            <li>Maintain discipline and behave professionally.</li>
            <li>If taking leave, inform in course WhatsApp group.</li>
            <li>More than 2 days leave requires proof + email.</li>
            <li>Placement assistance provided (not guarantee).</li>
        </ol>

        <p>Let's Learn & Grow Together 🚀</p>

        <br>

        <strong>Errors2Experts Team</strong>
        """

        qr_path = os.path.join(
            settings.BASE_DIR,
            "static",
            "images",
            "qr.jpeg"
        )

        send_brevo_email(
            email,
            "Course Booking Confirmation - Errors2Experts",
            student_html,
            attachment_path=qr_path,
        )

        # ---------------- ADMIN EMAIL ----------------

        admin_subject = "🚀 New Course Booking Received"

        admin_html = f"""
        <h2>New Course Booking</h2>

        <p><strong>Name:</strong> {name}</p>
        <p><strong>Email:</strong> {email}</p>
        <p><strong>Mobile:</strong> {mobile}</p>
        <p><strong>Education:</strong> {education}</p>
        <p><strong>Year Passed:</strong> {year_passed}</p>

        <hr>

        <p><strong>Course:</strong> {course.title}</p>
        <p><strong>Payment Type:</strong> {payment_type}</p>
        <p><strong>First Payment:</strong> ₹{amount}</p>
        <p><strong>Balance:</strong> ₹{balance}</p>
        <p><strong>Due Date:</strong> {due_date.strftime('%d-%m-%Y') if due_date else 'N/A'}</p>

        <hr>

        <p><strong>Booking Time:</strong> {datetime.now().strftime('%d-%m-%Y %H:%M')}</p>

        <p>Please follow up with the student.</p>
        """

        send_brevo_email(
            settings.ADMIN_NOTIFICATION_EMAIL,
            admin_subject,
            admin_html,
        )

        # ---------------- SHOW QR PAGE ----------------

        return render(request, "show_qr.html", {
            "amount": amount,
            "balance": balance,
            "due_date": due_date,
            "payment_type": payment_type,
            "course_title": course.title,
        })

    return redirect("course_detail", id=id)


def service_booking(request):
    if request.method == "POST":

        full_name = request.POST.get("full_name")
        email = request.POST.get("email")
        mobile = request.POST.get("mobile")
        service_name = request.POST.get("service_name")
        message = request.POST.get("message")
        preferred_date = request.POST.get("preferred_date")

        # ---------------- SAVE TO DATABASE ----------------

        ServiceBooking.objects.create(
            full_name=full_name,
            email=email,
            mobile=mobile,
            service_name=service_name,
            message=message,
            preferred_date=preferred_date,
        )

        # ---------------- SAVE TO EXCEL ----------------

        file_path = os.path.join(settings.BASE_DIR, "service_bookings.xlsx")

        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append([
                "Full Name",
                "Email",
                "Mobile",
                "Service",
                "Preferred Date",
                "Message",
            ])

        ws.append([
            full_name,
            email,
            mobile,
            service_name,
            preferred_date,
            message,
        ])

        wb.save(file_path)

        # ---------------- EMAIL TO ADMIN ----------------

        admin_subject = "🚀 New Service Booking - Errors2Experts"

        admin_html = f"""
        <h2>New Service Booking Received</h2>

        <table cellpadding="6">
            <tr>
                <td><strong>Name</strong></td>
                <td>{full_name}</td>
            </tr>
            <tr>
                <td><strong>Email</strong></td>
                <td>{email}</td>
            </tr>
            <tr>
                <td><strong>Mobile</strong></td>
                <td>{mobile}</td>
            </tr>
            <tr>
                <td><strong>Service</strong></td>
                <td>{service_name}</td>
            </tr>
            <tr>
                <td><strong>Preferred Date</strong></td>
                <td>{preferred_date}</td>
            </tr>
        </table>

        <hr>

        <h4>Customer Message</h4>

        <p>{message}</p>
        """

        send_brevo_email(
            settings.ADMIN_NOTIFICATION_EMAIL,
            admin_subject,
            admin_html,
        )

        # ---------------- EMAIL TO CLIENT ----------------

        client_subject = "Thank You for Registering"

        client_html = f"""
        <h2>Hello {full_name},</h2>

        <p>
        Thank you for registering for our
        <strong>{service_name}</strong> service.
        </p>

        <p>
        We have successfully received your request.
        </p>

        <p>
        Our team will contact you as soon as possible.
        </p>

        <hr>

        <h3>Stay Connected</h3>

        <p>
        🌐 Website: https://yourwebsite.com
        </p>

        <p>
        📸 Instagram:<br>
        https://www.instagram.com/errors2experts_2026/
        </p>

        <br>

        <p>
        Let's learn, grow, and build your future together!
        </p>

        <br>

        <strong>Best Regards,</strong><br>
        <strong>Errors2Experts Team</strong>
        """

        send_brevo_email(
            email,
            client_subject,
            client_html,
        )

        messages.success(request, "Service Registered Successfully")
        return redirect("services")

    return redirect("services")

# ---------------- WORKSHOP GALLERY ----------------


def workshop_gallery(request):
    workshops = WorkshopPhoto.objects.all()
    upcoming_workshop = UpcomingWorkshop.objects.order_by("event_date").first()

    return render(request, "workshop.html", {
        "workshops": workshops,
        "upcoming_workshop": upcoming_workshop,
    })


# ---------------- CERTIFICATE GALLERY ----------------

def certificate_gallery(request):
    certificates = Certificate.objects.all()

    return render(request, "certificate.html", {
        "certificates": certificates,
    })


# ---------------- INTERNSHIP LIST ----------------

def internship_list(request):
    internships = Internship.objects.all()

    return render(request, "all_internship.html", {
        "internships": internships,
    })


# ---------------- INTERNSHIP DETAIL ----------------

def internship_detail(request, pk):
    internship = get_object_or_404(Internship, pk=pk)

    timeline_data = defaultdict(list)

    for index, topic in enumerate(internship.syllabus):
        week = (index // 5) + 1
        timeline_data[week].append({
            "day": index + 1,
            "topic": topic,
        })

    return render(request, "internship_detail.html", {
        "internship": internship,
        "timeline_data": dict(timeline_data),
    })


def workshop_registration(request):

    if request.method == "POST":

        # ================= GET FORM DATA =================

        full_name = request.POST.get("full_name")
        email = request.POST.get("email")
        mobile = request.POST.get("mobile")
        gender = request.POST.get("gender")

        participation_mode = request.POST.get("participation_mode")
        present_address = request.POST.get("present_address")
        year_of_study = request.POST.get("year_of_study")
        college_name = request.POST.get("college_name")
        degree = request.POST.get("degree")
        department = request.POST.get("department")
        passed_out_year = request.POST.get("passed_out_year")

        interested_domains = ", ".join(
            request.POST.getlist("interested_domains")
        )

        technical_skill = request.POST.get("technical_skill")
        technical_skill_other = request.POST.get("technical_skill_other")

        demo_interest = request.POST.get("demo_interest")
        demo_interest_other = request.POST.get("demo_interest_other")

        demo_mode = request.POST.get("demo_mode")

        preferred_demo_time = ", ".join(
            request.POST.getlist("preferred_demo_time")
        )

        consent = request.POST.get("consent") == "True"

        queries = request.POST.get("queries")
        referred_by = request.POST.get("referred_by")

        # ================= DATABASE =================

        WorkshopRegistration.objects.create(
            full_name=full_name,
            email=email,
            mobile=mobile,
            gender=gender,
            participation_mode=participation_mode,
            present_address=present_address,
            year_of_study=year_of_study,
            college_name=college_name,
            degree=degree,
            department=department,
            passed_out_year=passed_out_year,
            interested_domains=interested_domains,
            technical_skill=technical_skill,
            technical_skill_other=technical_skill_other,
            demo_interest=demo_interest,
            demo_interest_other=demo_interest_other,
            demo_mode=demo_mode,
            preferred_demo_time=preferred_demo_time,
            consent=consent,
            queries=queries,
            referred_by=referred_by,
        )

        # ================= SAVE TO EXCEL =================

        file_path = os.path.join(settings.BASE_DIR, "workshop_registrations.xlsx")

        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active

        else:
            wb = Workbook()
            ws = wb.active

            # Header Row
            ws.append([
                "Full Name","Email","Mobile","Gender","Participation Mode", "Present Address","Year Of Study",
                "College Name","Degree","Department","Passed Out Year","Interested Domains","Technical Skill",
                "Technical Skill (Other)","Demo Interest",
                "Demo Interest (Other)",
                "Preferred Demo Mode",
                "Preferred Demo Time",
                "Consent",
                "Queries",
                "Referred By",
                "Registration Date"
            ])
            ws.append([
                "Full Name",
                "Email",
                "Mobile",
                "Gender",
                "Participation Mode",
                "Present Address",
                "Year Of Study",
                "College Name",
                "Degree",
                "Department",
                "Passed Out Year",
                "Interested Domains",
                "Technical Skill",
                "Technical Skill (Other)",
                "Demo Interest",
                "Demo Interest (Other)",
                "Preferred Demo Mode",
                "Preferred Demo Time",
                "Consent",
                "Queries",
                "Referred By",
                "Registration Date"
            ])

            # Data Row
            ws.append([
                full_name,
                email,
                mobile,
                gender,
                participation_mode,
                present_address,
                year_of_study,
                college_name,
                degree,
                department,
                passed_out_year,
                interested_domains,
                technical_skill,
                technical_skill_other,
                demo_interest,
                demo_interest_other,
                demo_mode,
                preferred_demo_time,
                "Yes" if consent else "No",
                queries,
                referred_by,
                datetime.now().strftime("%d-%m-%Y %I:%M %p")
            ])

            wb.save(file_path)

            # =====================================================
            # ADMIN EMAIL
            # =====================================================

            admin_subject = f"📢 New Workshop Registration - {full_name}"

            admin_html = f"""
            <!DOCTYPE html>
            <html>

            <body style="font-family:Arial,Helvetica,sans-serif;background:#f4f4f4;padding:30px;">

            <div style="
                max-width:800px;
                margin:auto;
                background:#ffffff;
                border-radius:10px;
                overflow:hidden;
                border:1px solid #dddddd;
            ">

                <div style="background:#2e7d32;padding:18px;color:white;">
                    <h2 style="margin:0;">
                        📢 New Workshop Registration
                    </h2>
                </div>

                <div style="padding:25px;">

                    <table style="width:100%;border-collapse:collapse;">

                        <tr>
                            <td style="padding:10px;border:1px solid #ddd;"><b>Full Name</b></td>
                            <td style="padding:10px;border:1px solid #ddd;">{full_name}</td>
                        </tr>

                        <tr>
                            <td style="padding:10px;border:1px solid #ddd;"><b>Email</b></td>
                            <td style="padding:10px;border:1px solid #ddd;">{email}</td>
                        </tr>

                        <tr>
                            <td style="padding:10px;border:1px solid #ddd;"><b>Mobile</b></td>
                            <td style="padding:10px;border:1px solid #ddd;">{mobile}</td>
                        </tr>

                        <tr>
                            <td style="padding:10px;border:1px solid #ddd;"><b>Gender</b></td>
                            <td style="padding:10px;border:1px solid #ddd;">{gender}</td>
                        </tr>

                        <tr>
                            <td style="padding:10px;border:1px solid #ddd;"><b>Participation Mode</b></td>
                            <td style="padding:10px;border:1px solid #ddd;">{participation_mode}</td>
                        </tr>

                        <tr>
                            <td style="padding:10px;border:1px solid #ddd;"><b>Present Address</b></td>
                            <td style="padding:10px;border:1px solid #ddd;">{present_address}</td>
                        </tr>

                        <tr>
                            <td style="padding:10px;border:1px solid #ddd;"><b>Year Of Study</b></td>
                            <td style="padding:10px;border:1px solid #ddd;">{year_of_study}</td>
                        </tr>

                        <tr>
                            <td style="padding:10px;border:1px solid #ddd;"><b>College Name</b></td>
                            <td style="padding:10px;border:1px solid #ddd;">{college_name}</td>
                        </tr>

                        <tr>
                            <td style="padding:10px;border:1px solid #ddd;"><b>Degree</b></td>
                            <td style="padding:10px;border:1px solid #ddd;">{degree}</td>
                        </tr>

                        <tr>
                            <td style="padding:10px;border:1px solid #ddd;"><b>Department</b></td>
                            <td style="padding:10px;border:1px solid #ddd;">{department}</td>
                        </tr>

                        <tr>
                            <td style="padding:10px;border:1px solid #ddd;"><b>Passed Out Year</b></td>
                            <td style="padding:10px;border:1px solid #ddd;">{passed_out_year}</td>
                        </tr>

                        <tr>
                            <td style="padding:10px;border:1px solid #ddd;"><b>Interested Domains</b></td>
                            <td style="padding:10px;border:1px solid #ddd;">{interested_domains}</td>
                        </tr>

                        <tr>
                            <td style="padding:10px;border:1px solid #ddd;"><b>Technical Skill</b></td>
                            <td style="padding:10px;border:1px solid #ddd;">{technical_skill}</td>
                        </tr>

                        <tr>
                            <td style="padding:10px;border:1px solid #ddd;"><b>Technical Skill (Other)</b></td>
                            <td style="padding:10px;border:1px solid #ddd;">{technical_skill_other}</td>
                        </tr>

                        <tr>
                            <td style="padding:10px;border:1px solid #ddd;"><b>Demo Interest</b></td>
                            <td style="padding:10px;border:1px solid #ddd;">{demo_interest}</td>
                        </tr>

                        <tr>
                            <td style="padding:10px;border:1px solid #ddd;"><b>Demo Interest (Other)</b></td>
                            <td style="padding:10px;border:1px solid #ddd;">{demo_interest_other}</td>
                        </tr>

                        <tr>
                            <td style="padding:10px;border:1px solid #ddd;"><b>Preferred Demo Mode</b></td>
                            <td style="padding:10px;border:1px solid #ddd;">{demo_mode}</td>
                        </tr>

                        <tr>
                            <td style="padding:10px;border:1px solid #ddd;"><b>Preferred Demo Time</b></td>
                            <td style="padding:10px;border:1px solid #ddd;">{preferred_demo_time}</td>
                        </tr>

                        <tr>
                            <td style="padding:10px;border:1px solid #ddd;"><b>Consent</b></td>
                            <td style="padding:10px;border:1px solid #ddd;">{"Yes" if consent else "No"}</td>
                        </tr>

                        <tr>
                            <td style="padding:10px;border:1px solid #ddd;"><b>Queries</b></td>
                            <td style="padding:10px;border:1px solid #ddd;">{queries}</td>
                        </tr>

                        <tr>
                            <td style="padding:10px;border:1px solid #ddd;"><b>Referred By</b></td>
                            <td style="padding:10px;border:1px solid #ddd;">{referred_by}</td>
                        </tr>

                        <tr>
                            <td style="padding:10px;border:1px solid #ddd;"><b>Registration Time</b></td>
                            <td style="padding:10px;border:1px solid #ddd;">{datetime.now().strftime("%d-%m-%Y %I:%M %p")}</td>
                        </tr>

                    </table>

                </div>

            </div>

            </body>
            </html>
            """

            send_brevo_email(
                settings.ADMIN_NOTIFICATION_EMAIL,
                admin_subject,
                admin_html
            )

            # =====================================================
            # USER EMAIL
            # =====================================================

            user_subject = "Workshop Registration Successful | Errors2Experts"

            user_html = f"""
            <!DOCTYPE html>
            <html>

            <body style="margin:0;padding:0;background:#f4f6f9;font-family:Arial,Helvetica,sans-serif;">

            <div style="max-width:700px;margin:30px auto;background:#ffffff;border-radius:12px;overflow:hidden;border:1px solid #e5e5e5;">

                <!-- Header -->
                <div style="background:#2e7d32;padding:25px;text-align:center;color:white;">
                    <h2 style="margin:0;">🎉 Registration Successful</h2>
                    <p style="margin-top:8px;font-size:15px;">
                        Thank you for registering with Errors2Experts.
                    </p>
                </div>

                <!-- Body -->
                <div style="padding:30px;">

                    <p>Dear <strong>{full_name}</strong>,</p>

                    <p>
                        Thank you for registering for our training program.
                        We have successfully received your registration details.
                    </p>

                    <p>
                        Our team will contact you shortly regarding your preferred
                        domain and demo session (if selected).
                    </p>

                    <h3 style="color:#2e7d32;border-bottom:2px solid #2e7d32;padding-bottom:8px;">
                        Registration Summary
                    </h3>

                    <table style="width:100%;border-collapse:collapse;">

                        <tr>
                            <td style="padding:8px;border:1px solid #ddd;"><b>Name</b></td>
                            <td style="padding:8px;border:1px solid #ddd;">{full_name}</td>
                        </tr>

                        <tr>
                            <td style="padding:8px;border:1px solid #ddd;"><b>Email</b></td>
                            <td style="padding:8px;border:1px solid #ddd;">{email}</td>
                        </tr>

                        <tr>
                            <td style="padding:8px;border:1px solid #ddd;"><b>Mobile</b></td>
                            <td style="padding:8px;border:1px solid #ddd;">{mobile}</td>
                        </tr>

                        <tr>
                            <td style="padding:8px;border:1px solid #ddd;"><b>Participation Mode</b></td>
                            <td style="padding:8px;border:1px solid #ddd;">{participation_mode}</td>
                        </tr>

                        <tr>
                            <td style="padding:8px;border:1px solid #ddd;"><b>Interested Domains</b></td>
                            <td style="padding:8px;border:1px solid #ddd;">{interested_domains}</td>
                        </tr>

                        <tr>
                            <td style="padding:8px;border:1px solid #ddd;"><b>Technical Skill</b></td>
                            <td style="padding:8px;border:1px solid #ddd;">{technical_skill}</td>
                        </tr>

                        <tr>
                            <td style="padding:8px;border:1px solid #ddd;"><b>Demo Interest</b></td>
                            <td style="padding:8px;border:1px solid #ddd;">{demo_interest}</td>
                        </tr>

                        <tr>
                            <td style="padding:8px;border:1px solid #ddd;"><b>Preferred Demo Mode</b></td>
                            <td style="padding:8px;border:1px solid #ddd;">{demo_mode}</td>
                        </tr>

                        <tr>
                            <td style="padding:8px;border:1px solid #ddd;"><b>Preferred Demo Time</b></td>
                            <td style="padding:8px;border:1px solid #ddd;">{preferred_demo_time}</td>
                        </tr>

                    </table>

                    <br>

                    <div style="background:#f8f9fa;padding:15px;border-left:4px solid #2e7d32;">
                        <strong>What's Next?</strong>
                        <ul style="margin:10px 0 0 18px;padding:0;">
                            <li>Our team will review your registration.</li>
                            <li>You'll receive a call or email regarding the next steps.</li>
                            <li>If you've requested a demo session, we'll schedule it based on your preferred mode and time.</li>
                        </ul>
                    </div>

                    <br>

                    <p>
                        If you have any questions, simply reply to this email or contact our support team.
                    </p>

                    <p>
                        Regards,<br>
                        <strong>Errors2Experts Team</strong>
                    </p>

                </div>

                <!-- Footer -->
                <div style="background:#f1f1f1;padding:15px;text-align:center;font-size:13px;color:#666;">
                    © 2026 Errors2Experts. All Rights Reserved.
                </div>

            </div>

            </body>
            </html>
            """

            send_brevo_email(
                email,
                user_subject,
                user_html
            )

        
        return redirect("home")

    return redirect("home")

def demo_request(request):
    if request.method == "POST":

        organization_name = request.POST.get("organization_name")
        email = request.POST.get("email")
        mobile = request.POST.get("mobile")
        category_value = request.POST.get("category")
        custom_requirement = request.POST.get("custom_requirement")

        # ================= GET CATEGORY =================

        if category_value == "other":
            category = None
        else:
            try:
                category = DemoCategory.objects.get(id=category_value)
            except DemoCategory.DoesNotExist:
                category = None

        # ================= SAVE REQUEST =================

        DemoRequest.objects.create(
            organization_name=organization_name,
            email=email,
            mobile=mobile,
            category=category,
            custom_requirement=custom_requirement,
        )

        # ===================================================
        # CATEGORY DEMO REQUEST
        # ===================================================

        if category:

            user_subject = f"{category.name} Demo Link"

            user_html = f"""
            <h2>Hello {organization_name},</h2>

            <p>
            Thank you for requesting the
            <b>{category.name}</b> demo.
            </p>

            <p>
            Below is your demo link:
            </p>

            <p>
            <a href="{category.demo_link}">
                {category.demo_link}
            </a>
            </p>

            <br>

            <p>
            Thank you.
            </p>

            <p>
            Regards,<br>
            <b>Errors2Experts Team</b>
            </p>
            """

            send_brevo_email(
                email,
                user_subject,
                user_html
            )

            messages.success(
                request,
                "Demo link has been sent to your email."
            )

        # ===================================================
        # CUSTOM REQUIREMENT
        # ===================================================

        else:

            # ---------------- USER EMAIL ----------------

            user_subject = "Demo Request Received"

            user_html = f"""
            <h2>Hello {organization_name},</h2>

            <p>
            Thank you for contacting
            <b>Errors2Experts</b>.
            </p>

            <p>
            We have received your custom demo request.
            </p>

            <p>
            Our technical team will review your requirement.
            </p>

            <p>
            You will receive a response within
            <b>2-3 business days.</b>
            </p>

            <br>

            <p>
            Regards,<br>
            <b>Errors2Experts Team</b>
            </p>
            """

            send_brevo_email(
                email,
                user_subject,
                user_html
            )

            # ---------------- ADMIN EMAIL ----------------

            admin_subject = "🚨 Critical Custom Demo Request"

            admin_html = f"""
            <h2>🚨 New Custom Demo Request</h2>

            <table cellpadding="6">
                <tr>
                    <td><b>Customer Name</b></td>
                    <td>{organization_name}</td>
                </tr>

                <tr>
                    <td><b>Email</b></td>
                    <td>{email}</td>
                </tr>

                <tr>
                    <td><b>Mobile</b></td>
                    <td>{mobile}</td>
                </tr>
            </table>

            <br>

            <h3>Requirement</h3>

            <p>{custom_requirement}</p>

            <br>

            <b>Errors2Experts System</b>
            """

            send_brevo_email(
                settings.ADMIN_NOTIFICATION_EMAIL,
                admin_subject,
                admin_html
            )

            messages.success(
                request,
                "Your request has been received successfully."
            )

        return redirect("services")

    return redirect("services")
    

