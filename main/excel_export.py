from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


def export_as_excel(modeladmin, request, queryset):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = modeladmin.model._meta.verbose_name_plural.title()

    fields = [field for field in modeladmin.model._meta.fields if field.name != "id"]

    # Header Style
    header_fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )

    header_font = Font(bold=True, color="FFFFFF")

    for col_num, field in enumerate(fields, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = field.verbose_name.title()
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(fields, 1):
            value = getattr(obj, field.name)

            if value is None:
                value = ""
            elif hasattr(value, "strftime"):
                value = value.strftime("%d-%m-%Y %I:%M %p")
            else:
                value = str(value)

            worksheet.cell(row=row_num, column=col_num, value=value)

    # Auto Width
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = (
            length + 5
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        f'attachment; filename="{modeladmin.model._meta.model_name}.xlsx"'
    )

    workbook.save(response)

    return response


export_as_excel.short_description = "📥 Export Selected to Excel"
